from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import openpyxl
import time

driver = webdriver.Chrome('chromedriver.exe')

driver.get('https://comic.naver.com/webtoon/weekdayList?week=sun')
driver.implicitly_wait(5)

WB = openpyxl.Workbook()
WB.create_sheet("naver")  # 시트 추가
sheet = WB["naver"]
sheet['A1'] = "id"
sheet["B1"] = "title"

for i in range(3, 100):
    try:
        title = driver.find_elements_by_css_selector(
            ".thumb a")[i].get_attribute("title")
        sheet.cell(i+2, 1, i)
        sheet.cell(i+2, 2, title)
    except:
        print("Error!")
        WB.save("NaverDay.xls")
        break

WB.save("NaverDay.xls")
print("DONE!")
