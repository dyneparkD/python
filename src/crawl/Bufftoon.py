from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import openpyxl
from selenium.webdriver.common.keys import Keys
import time

chrome_driver = "C:\\Users\\s6xya\\Desktop\\dyneparkD\\python\\chromedriver.exe"
chrome_options = Options()
chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:8989")
driver = webdriver.Chrome(chrome_driver, chrome_options=chrome_options)

#####################################################
WB = openpyxl.Workbook()  # 엑셀 파일 생성
WB.create_sheet("bufftoon")  # 시트 추가
sheet = WB["bufftoon"]  # 시트 선택
sheet['A1'] = "id"
sheet["B1"] = "title"
sheet["C1"] = "author"
sheet["D1"] = "synopsis"
sheet["E1"] = "genre"
sheet["F1"] = "likes"
sheet["G1"] = "image"
sheet["H1"] = "url"
sheet["I1"] = "views"
sheet["J1"] = "keywords"
#####################################################


driver.get('https://bufftoon.plaync.com/tag/original?currentType=webtoon')
driver.implicitly_wait(5)
print("bufftoon original")
driver.implicitly_wait(5)
time.sleep(0.5)
html = driver.find_element_by_tag_name('html')


def scrollDown():
    scrolled = 0
    SCROLL_PAUSE_TIME = 0.5
    # *change max_height
    max_height = 16300
    last_height = driver.execute_script("return document.body.scrollHeight")

    while last_height <= max_height:
        driver.execute_script(
            "window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(SCROLL_PAUSE_TIME)
        new_height = driver.execute_script("return document.body.scrollHeight")
        print("scrolling..." + str(scrolled))
        scrolled = scrolled + 1
        if new_height == last_height:
            scrolled = 0
            break
        last_height = new_height


for i in range(1, 161):
    try:
        scrollDown()
        driver.find_element_by_xpath(
            f"//*[@id='content']/div[2]/div/ul/li[{i}]").click()
        time.sleep(1)
        url = driver.current_url
        image = driver.find_element_by_xpath(
            "//*[@id='content']/div[2]/div/div[1]/div[1]/div[1]/span/img").get_attribute("src")
        title = driver.find_element_by_class_name("title").text
        author = driver.find_element_by_class_name("author").text
        genre = driver.find_element_by_class_name("genre").text
        views = driver.find_element_by_class_name("page-view-count").text
        keywords = driver.find_element_by_css_selector(".description p").text

        sheet['A1'] = "id"
        sheet["B1"] = "title"
        sheet["C1"] = "author"
        sheet["D1"] = "synopsis"
        sheet["E1"] = "genre"
        sheet["F1"] = "likes"
        sheet["G1"] = "image"
        sheet["H1"] = "url"
        sheet["I1"] = "views"
        sheet["J1"] = "keywords"

        sheet.cell(i+2, 1, i)
        sheet.cell(i+2, 2, title)
        sheet.cell(i+2, 3, author)
        sheet.cell(i+2, 5, genre)
        sheet.cell(i+2, 7, image)
        sheet.cell(i+2, 8, url)
        sheet.cell(i+2, 9, views)
        sheet.cell(i+2, 10, keywords)

        print(title)
        print(i)
        driver.get('https://bufftoon.plaync.com/tag/original?currentType=webtoon')
        time.sleep(0.5)
        driver.implicitly_wait(5)
        scrollDown()
    except:
        print("Error!")
        WB.save("bufftoon.xls")
        break

WB.save("bufftoon.xls")
print("DONE!")

# synopsis = driver.find_element_by_xpath(
#     "//meta[@name='description']").get_attribute("content")
