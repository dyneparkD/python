from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from time import sleep
import openpyxl

WB = openpyxl.Workbook()  # 엑셀 파일 생성
WB.create_sheet("Naver")  # 시트 추가
sheet = WB["Naver"]  # 시트 선택
sheet['A1'] = "id"
sheet["B1"] = "title"
sheet["C1"] = "author"
sheet["D1"] = "synopsis"
sheet["E1"] = "genre"
sheet["F1"] = "age"
sheet["G1"] = "likes"
sheet["H1"] = "image"
sheet["I1"] = "url"


# 웹페이지 접속
driver = webdriver.Chrome('chromedriver.exe')
driver.get('https://comic.naver.com/webtoon/artist')
driver.implicitly_wait(5)

count = 0
indexError = []
webtoons = driver.find_elements_by_class_name('thumb a')

for webtoon in webtoons:
    x = driver.find_elements_by_class_name(
        'thumb a')[count].get_attribute('href')
    driver.get(x)
    driver.implicitly_wait(5)

    title = driver.find_elements_by_css_selector(".detail h2")[0].text
    author = driver.find_elements_by_css_selector(".wrt_nm")[0].text
    synopsis = driver.find_elements_by_css_selector(".detail p")[0].text
    genre = driver.find_elements_by_css_selector(".genre")[0].text
    try:
        age = driver.find_elements_by_css_selector(".age")[0].text
    except IndexError:
        age = "null"
        print("-----error-----")
        indexError.append(count)
        print(indexError)
        print("-----error-----")
    try:
        likes = driver.find_elements_by_css_selector(".u_cnt")[0].text
    except IndexError:
        likes = "null"
        print("-----error-----")
        indexError.append(count)
        print(indexError)
        print("-----error-----")
    image = driver.find_element_by_css_selector(
        ".thumb a img").get_attribute("src")
    url = driver.current_url

    sheet.cell(count+2, 1, count)
    sheet.cell(count+2, 2, title)
    sheet.cell(count+2, 3, author)
    sheet.cell(count+2, 4, synopsis)
    sheet.cell(count+2, 5, genre)
    sheet.cell(count+2, 6, age)
    sheet.cell(count+2, 7, likes)
    sheet.cell(count+2, 8, image)
    sheet.cell(count+2, 9, url)

    count = count + 1
    print(count)
    print(len(webtoons))

    driver.back()
    driver.implicitly_wait(5)

WB.save("Test.xls")
print("DONE!")
# 네이버 = 2055개 웹툰
