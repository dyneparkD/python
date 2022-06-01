from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import openpyxl
import time

chrome_driver = "C:\\Users\\s6xya\\Desktop\\dyneparkD\\python\\chromedriver.exe"
chrome_options = Options()
chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:8989")
driver = webdriver.Chrome(chrome_driver, chrome_options=chrome_options)

# 웹툰원작
driver.get('https://webtoon.kakao.com/original-webtoon?tab=sun')
# 소설원작
# driver.get('https://webtoon.kakao.com/original-novel?tab=mon')
driver.implicitly_wait(5)
print("kakao")

#####################################################
# 카카오 페이지는 차단당함: 20개 크롤링 하니 403에러 나옴
#####################################################
WB = openpyxl.Workbook()  # 엑셀 파일 생성
WB.create_sheet("kakao")  # 시트 추가
sheet = WB["kakao"]  # 시트 선택
sheet['A1'] = "id"
sheet["B1"] = "title"
sheet["C1"] = "author"
sheet["D1"] = "synopsis"
sheet["E1"] = "genre"
sheet["F1"] = "likes"
sheet["G1"] = "image"
sheet["H1"] = "url"
sheet["I1"] = "views"
sheet["J1"] = "keywords"

#####################################################

# image = driver.find_elements_by_css_selector(".ContentCard_thumb__3HzZh")[
#     0].get_attribute("src")

print("sleeping~~~~~~~~~~~~~~~~~~~")

# day
count = 0
for i in range(0, 200):
    try:
        url = driver.find_elements_by_css_selector('.ContentCard_link__2B9Vi')[
            i].get_attribute('href')
        driver.get(url)
        driver.implicitly_wait(5)
        time.sleep(2)

        title = driver.find_element_by_xpath(
            "//meta[@name='og:title']").get_attribute("content")
        synopsis = driver.find_element_by_xpath(
            "//meta[@name='description']").get_attribute("content")
        keywords = driver.find_element_by_xpath(
            "//meta[@name='keywords']").get_attribute("content")
        image = driver.find_element_by_xpath(
            "//meta[@name='og:image']").get_attribute("content")
        genre = driver.find_elements_by_css_selector(
            ".Meta_countWrapper__1UNAH p")[0].text
        views = driver.find_elements_by_css_selector(
            ".Meta_countWrapper__1UNAH p")[1].text
        likes = driver.find_elements_by_css_selector(
            ".Meta_countWrapper__1UNAH p")[2].text

        sheet.cell(count+2, 1, count)
        sheet.cell(count+2, 2, title)
        sheet.cell(count+2, 4, synopsis)
        sheet.cell(count+2, 5, genre)
        sheet.cell(count+2, 6, likes)
        sheet.cell(count+2, 7, image)
        sheet.cell(count+2, 8, url)
        sheet.cell(count+2, 9, views)
        sheet.cell(count+2, 10, keywords)

        count = count + 1
        print(title)
        print(count)
        driver.get('https://webtoon.kakao.com/original-webtoon?tab=sun')
        time.sleep(2)
        driver.implicitly_wait(5)
    except:
        print("Error!")
        WB.save("kakao.xls")
        break

WB.save("kakao.xls")
print("DONE!")
